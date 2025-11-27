from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Lead
from .serializers import LeadSerializer
import csv
import io
import logging
from django.utils import timezone

logger = logging.getLogger(__name__)


class LeadListCreateView(generics.ListCreateAPIView):
    serializer_class = LeadSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Lead.objects.filter(user=self.request.user).order_by('-created_at')
        # Limit results if limit parameter is provided (for dashboard)
        limit = self.request.query_params.get('limit')
        if limit:
            try:
                queryset = queryset[:int(limit)]
            except ValueError:
                pass
        else:
            # Default limit of 500 for leads page to prevent slow loading
            queryset = queryset[:500]
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lead_stats(request):
    """Get lead statistics without fetching all leads"""
    total_leads = Lead.objects.filter(user=request.user).count()
    converted_leads = Lead.objects.filter(user=request.user, status='converted').count()
    qualified_leads = Lead.objects.filter(user=request.user, status='qualified').count()
    new_leads = Lead.objects.filter(user=request.user, status='new').count()
    
    return Response({
        'total': total_leads,
        'converted': converted_leads,
        'qualified': qualified_leads,
        'new': new_leads,
        'conversion_rate': round((converted_leads / total_leads * 100) if total_leads > 0 else 0, 2),
    }, status=status.HTTP_200_OK)


@api_view(['GET', 'PUT', 'DELETE'])
def lead_detail(request, pk):
    try:
        lead = Lead.objects.get(pk=pk, user=request.user)
    except Lead.DoesNotExist:
        return Response({'error': 'Lead not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = LeadSerializer(lead)
        return Response(serializer.data)
    elif request.method == 'PUT':
        serializer = LeadSerializer(lead, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'DELETE':
        lead.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_leads_csv(request):
    """
    Upload CSV or Excel file and create leads from it
    Expected CSV format:
    - Required columns: Name, Email
    - Optional columns: Phone, Source, Status, Service Type, Notes, Price, Potential Value, Description of Enquiry
    """
    logger.info(f"[Upload] Request received. Method: {request.method}, Content-Type: {request.content_type}")
    logger.info(f"[Upload] FILES keys: {list(request.FILES.keys()) if request.FILES else 'No FILES'}")
    
    if 'file' not in request.FILES:
        logger.error("[Upload] No file in request.FILES")
        return Response({'error': 'No file provided. Please select a file and try again.'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    logger.info(f"[Upload] File received: {file.name}, Size: {file.size}, Content-Type: {file.content_type}")
    
    # Check file extension
    if not file.name.endswith(('.csv', '.xlsx', '.xls')):
        logger.error(f"[Upload] Invalid file type: {file.name}")
        return Response({'error': 'Invalid file type. Please upload a CSV or Excel file (.csv, .xlsx, .xls)'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Read file based on type
        if file.name.endswith('.csv'):
            # Handle CSV
            decoded_file = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded_file))
            rows = list(csv_reader)
        elif file.name.endswith(('.xlsx', '.xls')):
            # Handle Excel files
            try:
                import openpyxl
                from openpyxl import load_workbook
            except ImportError:
                return Response({'error': 'Excel support requires openpyxl. Please install it or convert to CSV.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Read Excel file
            file.seek(0)  # Reset file pointer
            workbook = load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            # Read data rows
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        value = cell.value
                        # Convert to string, handle None
                        if value is None:
                            value = ''
                        else:
                            value = str(value).strip()
                        row_data[headers[idx]] = value
                # Only add row if it has at least one non-empty value
                if any(row_data.values()):
                    rows.append(row_data)
            
            # Convert to csv.DictReader-like format
            csv_reader = rows
        else:
            return Response({'error': 'Unsupported file type. Please upload CSV or Excel (.xlsx, .xls) file.'}, status=status.HTTP_400_BAD_REQUEST)
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # Column name mapping (handle different variations)
        column_mapping = {
            'name': ['Name', 'name', 'Full Name', 'full_name', 'FullName'],
            'email': ['Email', 'email', 'Email Address', 'email_address', 'EmailAddress'],
            'phone': ['Phone', 'phone', 'Phone Number', 'phone_number', 'PhoneNumber', 'Mobile', 'mobile'],
            'source': ['Source', 'source', 'Lead Source', 'lead_source', 'LeadSource'],
            'status': ['Status', 'status'],
            'service_type': ['Service Type', 'service_type', 'ServiceType', 'Service'],
            'notes': ['Notes', 'notes', 'Note', 'note', 'Comments', 'comments'],
            'price': ['Price', 'price'],
            'potential_value': ['Potential Value', 'potential_value', 'PotentialValue', 'Value'],
            'description_of_enquiry': ['Description of Enquiry', 'description_of_enquiry', 'Description', 'description', 'Enquiry', 'enquiry']
        }
        
        def find_column_value(row, possible_names):
            """Find value in row using possible column names"""
            for name in possible_names:
                if name in row and row[name]:
                    value = row[name]
                    # Handle both string and other types
                    if isinstance(value, str):
                        value = value.strip()
                    else:
                        value = str(value).strip() if value else ''
                    if value:
                        return value
            return None
        
        # Determine starting row number (2 for CSV with header, 2 for Excel with header)
        start_row = 2
        
        for row_num, row in enumerate(csv_reader, start=start_row):
            try:
                # Extract data using column mapping
                name = find_column_value(row, column_mapping['name'])
                email = find_column_value(row, column_mapping['email'])
                
                # Validate required fields
                if not name or not email:
                    skipped_count += 1
                    errors.append(f"Row {row_num}: Missing required fields (Name or Email)")
                    continue
                
                # Check if lead already exists (by email)
                existing_lead = Lead.objects.filter(user=request.user, email=email).first()
                
                # Extract optional fields
                phone = find_column_value(row, column_mapping['phone']) or ''
                source = find_column_value(row, column_mapping['source']) or 'Upload'
                status_value = find_column_value(row, column_mapping['status']) or 'new'
                service_type = find_column_value(row, column_mapping['service_type']) or ''
                notes = find_column_value(row, column_mapping['notes']) or ''
                
                # Parse numeric fields
                price = None
                price_str = find_column_value(row, column_mapping['price'])
                if price_str:
                    try:
                        price = float(price_str.replace(',', '').replace('$', '').strip())
                    except (ValueError, AttributeError):
                        pass
                
                potential_value = None
                value_str = find_column_value(row, column_mapping['potential_value'])
                if value_str:
                    try:
                        potential_value = float(value_str.replace(',', '').replace('$', '').strip())
                    except (ValueError, AttributeError):
                        pass
                
                description = find_column_value(row, column_mapping['description_of_enquiry']) or ''
                
                # Validate status
                valid_statuses = [choice[0] for choice in Lead.STATUS_CHOICES]
                if status_value not in valid_statuses:
                    status_value = 'new'
                
                # Validate service type
                valid_service_types = [choice[0] for choice in Lead.SERVICE_TYPE_CHOICES]
                if service_type and service_type not in valid_service_types:
                    service_type = ''
                
                if existing_lead:
                    # Update existing lead
                    existing_lead.name = name
                    existing_lead.phone = phone or existing_lead.phone
                    if source and source != 'Upload':
                        existing_lead.source = source
                    if status_value:
                        existing_lead.status = status_value
                    if service_type:
                        existing_lead.service_type = service_type
                    if notes:
                        existing_lead.notes = notes if not existing_lead.notes else f"{existing_lead.notes}\n{notes}"
                    if price is not None:
                        existing_lead.price = price
                    if potential_value is not None:
                        existing_lead.potential_value = potential_value
                    if description:
                        existing_lead.description_of_enquiry = description
                    existing_lead.save()
                    updated_count += 1
                else:
                    # Create new lead
                    lead = Lead.objects.create(
                        user=request.user,
                        name=name,
                        email=email,
                        phone=phone,
                        source=source,
                        status=status_value,
                        service_type=service_type,
                        notes=notes,
                        price=price,
                        potential_value=potential_value,
                        description_of_enquiry=description
                    )
                    created_count += 1
                    
                    # Trigger automation for new leads
                    try:
                        from automations.services import trigger_automations
                        trigger_automations('new_lead', {
                            'user': request.user,
                            'lead': lead
                        })
                    except Exception as e:
                        logger.error(f"Error triggering automation for lead {lead.id}: {str(e)}")
            
            except Exception as e:
                skipped_count += 1
                errors.append(f"Row {row_num}: {str(e)}")
                logger.error(f"Error processing row {row_num}: {str(e)}")
        
        return Response({
            'success': True,
            'message': f'Import completed: {created_count} created, {updated_count} updated, {skipped_count} skipped',
            'stats': {
                'total': created_count + updated_count + skipped_count,
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'errors_count': len(errors)
            },
            'errors': errors[:10]  # Return first 10 errors
        }, status=status.HTTP_200_OK)
    
    except Exception as e:
        logger.error(f"Error processing CSV file: {str(e)}")
        return Response({
            'error': f'Error processing file: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)
